# ============================================================
# routes_home.py — Home / Overview Page
# PSE Quant SaaS — Dashboard
# ============================================================

import sys
import os
import csv
import io
import json
from datetime import datetime
from pathlib import Path
from flask import Blueprint, render_template, jsonify, Response
from dashboard.security import rate_limit

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / 'db'))
sys.path.insert(0, str(ROOT))

import database as db
from db.db_settings import get_setting
from dashboard.db_members import (
    expire_overdue_members, get_member_stats,
    get_expiring_soon, get_recent_activity,
)
from dashboard.background import get_status as get_job_status

home_bp = Blueprint('home', __name__)


@home_bp.route('/')
def index():
    # Auto-expire overdue members on every page load
    expire_overdue_members()

    member_stats  = get_member_stats()
    expiring_soon = get_expiring_soon(days=7)
    activity      = get_recent_activity(limit=20)

    # Latest rankings from scores_v2 (dividend portfolio, top 10)
    scores_raw = db.get_last_scores_v2('dividend') or []
    unified_rankings = [
        {'ticker': s['ticker'], 'score': round(s['score'], 1)}
        for s in scores_raw[:10]
    ]

    # Pricing for quick payment link card
    pm_configured = bool(os.getenv('PAYMONGO_SECRET_KEY', ''))
    pm_monthly    = int(get_setting('monthly_price_centavos',
                                     os.getenv('MONTHLY_PRICE_CENTAVOS', 29900))) / 100
    pm_annual     = int(get_setting('annual_price_centavos',
                                     os.getenv('ANNUAL_PRICE_CENTAVOS', 299900))) / 100

    context = {
        'member_stats':      member_stats,
        'expiring_soon':     expiring_soon,
        'activity':          activity,
        'unified_rankings':  unified_rankings,
        'now':               datetime.now().strftime('%Y-%m-%d %H:%M'),
        'pm_configured':     pm_configured,
        'pm_monthly':        pm_monthly,
        'pm_annual':         pm_annual,
    }
    return render_template('home.html', **context)


@home_bp.route('/api/status')
def api_status():
    """JSON endpoint: system health snapshot."""
    try:
        ticker_count = len(db.get_all_tickers())
    except Exception:
        ticker_count = 0

    db_path = db.DB_PATH
    try:
        db_size_kb = round(os.path.getsize(db_path) / 1024, 1)
    except Exception:
        db_size_kb = 0

    member_stats = get_member_stats()
    job          = get_job_status()

    # Last score run date from DB
    try:
        conn = db.get_connection()
        row  = conn.execute(
            "SELECT MAX(run_date) AS run_date FROM scores_v2"
        ).fetchone()
        conn.close()
        last_run = row['run_date'] if row and row['run_date'] else 'Never'
    except Exception:
        last_run = 'Unknown'

    return jsonify({
        'stocks_tracked':  ticker_count,
        'db_size_kb':      db_size_kb,
        'last_score_run':  last_run,
        'active_members':  member_stats.get('active', 0),
        'expired_members': member_stats.get('expired', 0),
        'job':             job,
        'timestamp':       datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    })


@home_bp.route('/api/activity')
def api_activity():
    """JSON endpoint: recent activity log."""
    items = get_recent_activity(limit=30)
    return jsonify(items)


@home_bp.route('/api/rankings/export')
@rate_limit(limit=10)
def api_rankings_export():
    """CSV download of current unified rankings."""
    scores_raw = db.get_last_scores_v2('dividend') or []
    if not scores_raw:
        return jsonify({'error': 'No rankings available yet.'}), 404

    # Enrich with stock names
    conn = db.get_connection()
    name_map = {}
    try:
        rows = conn.execute("SELECT ticker, name, sector FROM stocks").fetchall()
        name_map = {r['ticker']: {'name': r['name'], 'sector': r['sector']} for r in rows}
    finally:
        conn.close()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Rank', 'Ticker', 'Company', 'Sector', 'Score'])
    for i, s in enumerate(sorted(scores_raw, key=lambda x: x.get('score', 0), reverse=True), 1):
        t = s['ticker']
        info = name_map.get(t, {})
        writer.writerow([i, t, info.get('name', ''), info.get('sector', ''), s.get('score', '')])

    today = datetime.now().strftime('%Y-%m-%d')
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename=pse_rankings_{today}.csv'},
    )


@home_bp.route('/export/rankings.csv')
def export_rankings_csv():
    """CSV download of current StockPilot PH Rankings (all portfolio types)."""
    try:
        rows = _build_export_rows()
        if not rows:
            return jsonify({'error': 'No rankings data available. Run the pipeline first.'}), 404

        output = io.StringIO()
        fieldnames = [
            'rank', 'ticker', 'name', 'sector', 'portfolio_type',
            'score', 'confidence', 'health_score', 'improvement_score',
            'persistence_score', 'run_date',
        ]
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

        today = datetime.now().strftime('%Y-%m-%d')
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename=stockpilot_rankings_{today}.csv'},
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@home_bp.route('/export/rankings.xlsx')
def export_rankings_xlsx():
    """Excel download of current StockPilot PH Rankings (4 sheets, one per portfolio type)."""
    if not EXCEL_AVAILABLE:
        return jsonify({'error': 'openpyxl not installed. Run: py -m pip install openpyxl'}), 200

    try:
        rows = _build_export_rows()
        if not rows:
            return jsonify({'error': 'No rankings data available. Run the pipeline first.'}), 404

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default empty sheet

        PORTFOLIO_TYPES = ['unified', 'dividend', 'value']
        HEADERS = [
            'Rank', 'Ticker', 'Name', 'Sector', 'Score', 'Confidence',
            'Health', 'Improvement', 'Persistence', 'Run Date',
        ]
        FIELD_KEYS = [
            'rank', 'ticker', 'name', 'sector', 'score', 'confidence',
            'health_score', 'improvement_score',
            'persistence_score', 'run_date',
        ]

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(fill_type='solid', fgColor='1F4E79')
        header_align = Alignment(horizontal='center', vertical='center')

        for pt in PORTFOLIO_TYPES:
            pt_rows = [r for r in rows if r['portfolio_type'] == pt]
            sheet_name = pt.replace('_', ' ').title()
            ws = wb.create_sheet(title=sheet_name)

            # Write header row
            ws.append(HEADERS)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

            # Freeze top row
            ws.freeze_panes = 'A2'

            # Write data rows
            for row in pt_rows:
                ws.append([row.get(k, '') for k in FIELD_KEYS])

            # Auto-fit column widths (approximate)
            col_widths = [6, 10, 35, 20, 8, 12, 10, 12, 14, 12]
            for i, width in enumerate(col_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        today = datetime.now().strftime('%Y-%m-%d')
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename=stockpilot_rankings_{today}.xlsx'},
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


def _build_export_rows() -> list:
    """
    Fetches scores_v2 for all portfolio types and builds flat export rows.
    Each row contains rank, ticker, name, sector, portfolio_type, score,
    confidence (as %), layer sub-scores, and run_date.
    """
    PORTFOLIO_TYPES = ['unified', 'dividend', 'value']

    # Fetch stock metadata (name, sector)
    conn = db.get_connection()
    name_map = {}
    try:
        meta_rows = conn.execute(
            "SELECT ticker, name, sector FROM stocks"
        ).fetchall()
        name_map = {r['ticker']: {'name': r['name'] or '', 'sector': r['sector'] or ''}
                    for r in meta_rows}
    finally:
        conn.close()

    all_rows = []
    for pt in PORTFOLIO_TYPES:
        # Get latest run_date for this portfolio type
        conn = db.get_connection()
        try:
            row = conn.execute(
                "SELECT MAX(run_date) AS latest FROM scores_v2 "
                "WHERE rank IS NOT NULL AND portfolio_type = ?",
                (pt,)
            ).fetchone()
            if not row or not row['latest']:
                continue
            latest = row['latest']

            # Fetch scores WITH breakdown_json
            score_rows = conn.execute(
                """SELECT ticker, score, rank, confidence, breakdown_json, run_date
                   FROM scores_v2
                   WHERE run_date = ? AND portfolio_type = ? AND rank IS NOT NULL
                   ORDER BY rank""",
                (latest, pt)
            ).fetchall()
        finally:
            conn.close()

        for sr in score_rows:
            ticker = sr['ticker']
            meta = name_map.get(ticker, {'name': '', 'sector': ''})

            # Parse breakdown JSON safely
            breakdown = {}
            if sr['breakdown_json']:
                try:
                    breakdown = json.loads(sr['breakdown_json'])
                except (ValueError, TypeError):
                    breakdown = {}

            layers  = breakdown.get('layers', {})
            health  = layers.get('health', {}).get('score')
            improve = layers.get('improvement', {}).get('score')
            persist = layers.get('persistence', {}).get('score')

            confidence_raw = sr['confidence']
            confidence_pct = (
                f"{round(confidence_raw * 100)}%"
                if confidence_raw is not None else ''
            )

            all_rows.append({
                'rank':               sr['rank'],
                'ticker':             ticker,
                'name':               meta['name'],
                'sector':             meta['sector'],
                'portfolio_type':     pt,
                'score':              round(sr['score'], 2) if sr['score'] is not None else '',
                'confidence':         confidence_pct,
                'health_score':       round(health, 2) if health is not None else '',
                'improvement_score':  round(improve, 2) if improve is not None else '',
                'persistence_score':  round(persist, 2) if persist is not None else '',
                'run_date':           sr['run_date'] or '',
            })

    return all_rows


@home_bp.route('/api/health')
def api_health():
    """Health check: DB connectivity, last run, disk, scheduler."""
    from dashboard.background import get_scheduler_status

    db_ok = False
    last_run = 'Never'
    ticker_count = 0
    db_size_kb = 0

    try:
        conn = db.get_connection()
        row = conn.execute("SELECT MAX(run_date) AS run_date FROM scores_v2").fetchone()
        conn.close()
        db_ok = True
        last_run = row['run_date'] if row and row['run_date'] else 'Never'
        ticker_count = len(db.get_all_tickers())
    except Exception as e:
        last_run = f'DB error: {e}'

    try:
        db_size_kb = round(os.path.getsize(db.DB_PATH) / 1024, 1)
    except Exception:
        pass

    sched = get_scheduler_status()

    return jsonify({
        'status':          'ok' if db_ok else 'degraded',
        'db_ok':           db_ok,
        'db_size_kb':      db_size_kb,
        'last_score_run':  last_run,
        'stocks_tracked':  ticker_count,
        'scheduler':       sched,
        'timestamp':       datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    })
